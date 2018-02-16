#!/usr/bin/env python

import csv
import openpyxl
import openpyxl.drawing
import os

from openpyxl.styles import Alignment

def go(options):

    with open(options.csv, 'r') as r:
        N = {row['netid']: row for row in csv.DictReader(r)}

    with open(options.csv0, 'r') as r:
        rows0 = list(csv.reader(r))

    if options.flipLR:
        for row in rows0:
            row.reverse()

    if options.flipTB:
        rows0.reverse()

    def join(X, s=None):
        def g():
            for x in X[:-1]:
                yield x
                yield s
            yield X[-1]
        return list(g())



    def g():
        def image(x):
            if x:
                path = os.path.join(options.netid, '{x}.png'.format(x=x))
                return openpyxl.drawing.image.Image(path)
            return None
        def name(x):
            if x:
                return '{last_name}, {first_name}'.format(**N[x])
            return None
        for row in rows0:
            yield join([image(x) for x in row])
            yield join([name(x) for x in row])
            yield join([x for x in row])
            yield join([None for x in row])

    rows1 = list(g())

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active

    for i, row in enumerate(rows1):
        for j, v in enumerate(row):
            c = ws1.cell(row=i+1, column=j+1)
            ws1.row_dimensions[c.row].height = 15
            ws1.column_dimensions[c.column].width = 2
        
    for i, row in enumerate(rows1):
        for j, v in enumerate(row):
            c = ws1.cell(row=i+1, column=j+1)
            if isinstance(v, openpyxl.drawing.image.Image):
                ws1.add_image(v, c.coordinate)
                ws1.row_dimensions[c.row].height = options.height
                ws1.column_dimensions[c.column].width = options.width
            else:
                c.alignment = Alignment(horizontal='center')
                c.value = v

    wb1.save(options.xlsx1)

import argparse

class HelpFormatter(argparse.ArgumentDefaultsHelpFormatter):
    def add_arguments(self, actions):
        actions = sorted(actions, key=attrgetter('option_strings'))
        super(HelpFormatter, self).add_arguments(actions)

def mkparser(description):
    parser = argparse.ArgumentParser(
                    description=description,
                    formatter_class=HelpFormatter,
                )
    parser.add_argument(
            '--debug', 
            type=int,
            default=0,
            help='''Turn on script debugging.'''
        )
    parser.add_argument(
            '--csv0', 
            required=True, 
            type=str,
            help='''The CSV input.'''
            )
    parser.add_argument(
            '--xlsx1', 
            required=True, 
            type=str,
            help='''The XLSX output.'''
            )
    parser.add_argument(
            '--netid', 
            required=True, 
            type=str,
            help='''The directory of NetIDs.'''
            )
    parser.add_argument(
            '--csv', 
            required=True, 
            type=str,
            help='''The CSV of known NetIDs.'''
            )
    parser.add_argument(
            '--width', 
            required=True, 
            type=int,
            help='''The width of the image-containing columns.'''
            )
    parser.add_argument(
            '--height', 
            required=True, 
            type=int,
            help='''The height of the image-containing columns.'''
            )
    parser.add_argument(
            '--flipLR', 
            default=False,
            action='store_true',
            help='''Flip left-to-right.'''
            )
    parser.add_argument(
            '--flipTB', 
            default=False,
            action='store_true',
            help='''Flip top-to-bottom.'''
            )
    return parser

if __name__ == '__main__':

    parser = mkparser('XLSX input to XLSX output.')

    options = parser.parse_args()

    go(options)


