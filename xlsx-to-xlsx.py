#!/usr/bin/env python

import csv
import openpyxl
import openpyxl.drawing
import os

from openpyxl.styles import Alignment

def go(options):

    with open(options.csv, 'r') as r:
        N = {row['netid']: row for row in csv.DictReader(r)}

    wb0 = openpyxl.load_workbook(options.xlsx0, read_only=False)
    ws0 = wb0.active

    def join(X, s=None):
        def g():
            for x in X[:-1]:
                yield x
                yield s
            yield X[-1]
        return list(g())

    def g():
        def image(x):
            if x is None:
                return None
            path = os.path.join(options.netid, '{x}.png'.format(x=x))
            return openpyxl.drawing.image.Image(path)
        def name(x):
            if x is None:
                return None
            return '{last_name}, {first_name}'.format(**N[x])
        for row in ws0.rows:
            yield join([image(x.value) for x in row])
            yield join([name(x.value) for x in row])
            yield join([x.value for x in row])
            yield join([None for x in row])

    rows = list(g())

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active

    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            c = ws1.cell(row=i+1, column=j+1)
            ws1.row_dimensions[c.row].height = 15
            ws1.column_dimensions[c.column].width = 2
        
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            c = ws1.cell(row=i+1, column=j+1)
            if isinstance(v, openpyxl.drawing.image.Image):
                ws1.add_image(v, c.coordinate)
                ws1.row_dimensions[c.row].height = options.height
                ws1.column_dimensions[c.column].width = options.width
            else:
                c.alignment = Alignment(horizontal='center')
                c.value = v

    wb1.save(options.xlsx1)

import argparse

class HelpFormatter(argparse.ArgumentDefaultsHelpFormatter):
    def add_arguments(self, actions):
        actions = sorted(actions, key=attrgetter('option_strings'))
        super(HelpFormatter, self).add_arguments(actions)

def mkparser(description):
    parser = argparse.ArgumentParser(
                    description=description,
                    formatter_class=HelpFormatter,
                )
    parser.add_argument(
            '--debug', 
            type=int,
            default=0,
            help='''Turn on script debugging.'''
        )
    parser.add_argument(
            '--xlsx0', 
            required=True, 
            type=str,
            help='''The XLSX input.'''
            )
    parser.add_argument(
            '--xlsx1', 
            required=True, 
            type=str,
            help='''The XLSX output.'''
            )
    parser.add_argument(
            '--netid', 
            required=True, 
            type=str,
            help='''The directory of NetIDs.'''
            )
    parser.add_argument(
            '--csv', 
            required=True, 
            type=str,
            help='''The CSV of known NetIDs.'''
            )
    parser.add_argument(
            '--width', 
            required=True, 
            type=int,
            help='''The width of the image-containing columns.'''
            )
    parser.add_argument(
            '--height', 
            required=True, 
            type=int,
            help='''The height of the image-containing columns.'''
            )
    return parser

if __name__ == '__main__':

    parser = mkparser('XLSX input to XLSX output.')

    options = parser.parse_args()

    go(options)


